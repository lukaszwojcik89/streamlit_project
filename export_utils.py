"""
Narzędzia do eksportu danych (Excel, CSV) dla aplikacji Raport Czasu Pracy.

Centralizuje logikę eksportu i stylowanie Excel, eliminując duplikaty z app.py.
"""

import io
from dataclasses import dataclass, field
from datetime import datetime
from typing import Dict, Optional, Any

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from config import EXCEL_COLORS, EXCEL_COLUMN_WIDTHS, get_color_for_percent
from helpers import hours_to_hm_format


# =============================================================================
# DATACLASS DLA STYLÓW EXCEL
# =============================================================================

@dataclass
class ExcelStyles:
    """
    Centralizuje wszystkie style Excel w jednym miejscu.

    Używane do formatowania nagłówków, danych i kolorowania komórek.
    """
    # Style nagłówka
    header_font: Font = field(default_factory=lambda: Font(
        name="Calibri", size=11, bold=True, color=EXCEL_COLORS["header_font"]
    ))
    header_fill: PatternFill = field(default_factory=lambda: PatternFill(
        start_color=EXCEL_COLORS["header_bg"],
        end_color=EXCEL_COLORS["header_bg"],
        fill_type="solid"
    ))
    header_alignment: Alignment = field(default_factory=lambda: Alignment(
        horizontal="center", vertical="center"
    ))

    # Style danych
    data_font: Font = field(default_factory=lambda: Font(name="Calibri", size=10))
    data_alignment_left: Alignment = field(default_factory=lambda: Alignment(
        horizontal="left", vertical="center"
    ))
    data_alignment_center: Alignment = field(default_factory=lambda: Alignment(
        horizontal="center", vertical="center"
    ))
    data_alignment_right: Alignment = field(default_factory=lambda: Alignment(
        horizontal="right", vertical="center"
    ))

    # Ramki
    thin_border: Border = field(default_factory=lambda: Border(
        left=Side(style="thin", color=EXCEL_COLORS["border"]),
        right=Side(style="thin", color=EXCEL_COLORS["border"]),
        top=Side(style="thin", color=EXCEL_COLORS["border"]),
        bottom=Side(style="thin", color=EXCEL_COLORS["border"]),
    ))

    # Kolory dla procentów twórczości
    color_red: PatternFill = field(default_factory=lambda: PatternFill(
        start_color=EXCEL_COLORS["red_fill"],
        end_color=EXCEL_COLORS["red_fill"],
        fill_type="solid"
    ))
    color_yellow: PatternFill = field(default_factory=lambda: PatternFill(
        start_color=EXCEL_COLORS["yellow_fill"],
        end_color=EXCEL_COLORS["yellow_fill"],
        fill_type="solid"
    ))
    color_green: PatternFill = field(default_factory=lambda: PatternFill(
        start_color=EXCEL_COLORS["green_fill"],
        end_color=EXCEL_COLORS["green_fill"],
        fill_type="solid"
    ))

    def get_fill_for_percent(self, percent: float) -> Optional[PatternFill]:
        """Zwraca odpowiedni fill dla procentu twórczości."""
        if pd.isna(percent):
            return None
        if percent <= 50:
            return self.color_red
        elif percent <= 80:
            return self.color_yellow
        else:
            return self.color_green


# =============================================================================
# FORMATOWANIE ARKUSZY EXCEL
# =============================================================================

def format_worksheet_headers(
    worksheet: Worksheet,
    styles: ExcelStyles,
    column_count: int
) -> None:
    """
    Formatuje nagłówki arkusza Excel.

    Args:
        worksheet: Arkusz do sformatowania
        styles: Obiekt ExcelStyles ze stylami
        column_count: Liczba kolumn do sformatowania
    """
    for cell in worksheet[1]:
        cell.font = styles.header_font
        cell.fill = styles.header_fill
        cell.alignment = styles.header_alignment
        cell.border = styles.thin_border


def format_worksheet_data(
    worksheet: Worksheet,
    styles: ExcelStyles,
    row_count: int,
    column_count: int,
    creative_percent_col: int = 6,
    numeric_cols: Dict[int, str] = None
) -> None:
    """
    Formatuje dane arkusza Excel z kolorowaniem procentów twórczości.

    Args:
        worksheet: Arkusz do sformatowania
        styles: Obiekt ExcelStyles ze stylami
        row_count: Liczba wierszy danych (bez nagłówka)
        column_count: Liczba kolumn
        creative_percent_col: Numer kolumny z procentem twórczości (1-based)
        numeric_cols: Dict {numer_kolumny: format_liczby} dla kolumn numerycznych
    """
    if numeric_cols is None:
        numeric_cols = {5: "0.00", 6: "0", 8: "0.00"}  # Domyślne: Czas(h), %, Godz.twórcze(h)

    for row_num in range(2, row_count + 2):
        # Pobierz wartość procentu do kolorowania
        creative_value = worksheet.cell(row=row_num, column=creative_percent_col).value

        for col_num in range(1, column_count + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.font = styles.data_font
            cell.border = styles.thin_border

            # Wyrównanie zależne od kolumny
            if col_num in [1, 2]:  # Osoba, Zadanie
                cell.alignment = styles.data_alignment_left
            elif col_num == 3:  # Klucz
                cell.alignment = styles.data_alignment_center
            else:  # Pozostałe (numeryczne)
                cell.alignment = styles.data_alignment_right

            # Format liczby jeśli określony
            if col_num in numeric_cols:
                cell.number_format = numeric_cols[col_num]

            # Kolorowanie procentu twórczości
            if col_num == creative_percent_col:
                if pd.notna(creative_value) and creative_value != "":
                    try:
                        percent = float(creative_value)
                        fill = styles.get_fill_for_percent(percent)
                        if fill:
                            cell.fill = fill
                    except (ValueError, TypeError):
                        pass


def set_column_widths(
    worksheet: Worksheet,
    widths: Dict[str, int] = None
) -> None:
    """
    Ustawia szerokości kolumn arkusza.

    Args:
        worksheet: Arkusz do sformatowania
        widths: Dict {litera_kolumny: szerokość}
    """
    if widths is None:
        widths = {
            "A": EXCEL_COLUMN_WIDTHS["osoba"],
            "B": EXCEL_COLUMN_WIDTHS["zadanie"],
            "C": EXCEL_COLUMN_WIDTHS["klucz"],
            "D": EXCEL_COLUMN_WIDTHS["czas"],
            "E": EXCEL_COLUMN_WIDTHS["czas_h"],
            "F": EXCEL_COLUMN_WIDTHS["procent"],
            "G": EXCEL_COLUMN_WIDTHS["godziny_tworcze"],
            "H": EXCEL_COLUMN_WIDTHS["godziny_tworcze_h"],
        }

    for col, width in widths.items():
        worksheet.column_dimensions[col].width = width


# =============================================================================
# GŁÓWNE FUNKCJE EKSPORTU
# =============================================================================

def export_to_csv(
    df: pd.DataFrame,
    columns: list,
    column_names: list,
    filename_prefix: str = "raport_pracy"
) -> tuple[str, str]:
    """
    Eksportuje DataFrame do CSV z polskimi znakami (UTF-8-BOM).

    Args:
        df: DataFrame do eksportu
        columns: Lista kolumn do wyeksportowania
        column_names: Nazwy kolumn w eksporcie
        filename_prefix: Prefix nazwy pliku

    Returns:
        Tuple (csv_data, filename)
    """
    export_df = df[columns].copy()
    export_df.columns = column_names

    csv_data = export_df.to_csv(index=False, encoding="utf-8-sig")
    filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    return csv_data, filename


def export_to_excel(
    df_detailed: pd.DataFrame,
    df_summary: pd.DataFrame,
    filename_prefix: str = "raport_pracy",
    sheet_name_detailed: str = "Raport pracy",
    sheet_name_summary: str = "Podsumowanie"
) -> tuple[io.BytesIO, str]:
    """
    Eksportuje dane do Excel z profesjonalnym formatowaniem (2 arkusze).

    Args:
        df_detailed: DataFrame ze szczegółowymi danymi
        df_summary: DataFrame z podsumowaniem per osoba
        filename_prefix: Prefix nazwy pliku
        sheet_name_detailed: Nazwa arkusza szczegółowego
        sheet_name_summary: Nazwa arkusza podsumowania

    Returns:
        Tuple (buffer, filename)
    """
    buffer = io.BytesIO()
    styles = ExcelStyles()

    # Przygotuj dane do eksportu Excel
    export_df = _prepare_excel_dataframe(df_detailed)

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Arkusz 1: Szczegółowe dane
        export_df.to_excel(writer, sheet_name=sheet_name_detailed, index=False)
        worksheet = writer.sheets[sheet_name_detailed]

        # Formatowanie
        set_column_widths(worksheet)
        format_worksheet_headers(worksheet, styles, len(export_df.columns))
        format_worksheet_data(
            worksheet, styles,
            row_count=len(export_df),
            column_count=len(export_df.columns)
        )

        # Zamrożenie i filtry
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:{chr(65 + len(export_df.columns) - 1)}{len(export_df) + 1}"

        # Arkusz 2: Podsumowanie
        summary_data = df_summary.reset_index() if df_summary.index.name else df_summary.copy()
        summary_data.to_excel(writer, sheet_name=sheet_name_summary, index=False)

        worksheet_summary = writer.sheets[sheet_name_summary]

        # Szerokości kolumn podsumowania
        summary_widths = {
            "A": 25,  # Osoba / index
            "B": 18,
            "C": 18,
            "D": 18,
            "E": 18,
        }
        set_column_widths(worksheet_summary, summary_widths)
        format_worksheet_headers(worksheet_summary, styles, len(summary_data.columns))

        # Formatuj dane podsumowania
        for row_num in range(2, len(summary_data) + 2):
            for col_num in range(1, len(summary_data.columns) + 1):
                cell = worksheet_summary.cell(row=row_num, column=col_num)
                cell.font = styles.data_font
                cell.border = styles.thin_border

                if col_num == 1:  # Osoba
                    cell.alignment = styles.data_alignment_left
                else:  # Liczby
                    cell.alignment = styles.data_alignment_right
                    if col_num in [2, 3]:  # Godziny
                        cell.number_format = "0.00"
                    else:  # Procenty
                        cell.number_format = "0.0"

        worksheet_summary.freeze_panes = "A2"

    buffer.seek(0)
    filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return buffer, filename


def _prepare_excel_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Przygotowuje DataFrame do eksportu Excel z odpowiednimi kolumnami.

    Args:
        df: DataFrame źródłowy z kolumnami: person, task, key, time_hours,
            creative_percent, creative_hours

    Returns:
        DataFrame gotowy do zapisu w Excel
    """
    export_df = pd.DataFrame()

    export_df["Osoba"] = df["person"]
    export_df["Zadanie"] = df["task"]
    export_df["Klucz"] = df["key"]
    export_df["Czas"] = df["time_hours"].apply(hours_to_hm_format)
    export_df["Czas (h)"] = df["time_hours"]
    export_df["Procent twórczości"] = df["creative_percent"]
    export_df["Godziny twórcze"] = df["creative_hours"].apply(
        lambda x: f"{x:.1f}h" if pd.notna(x) and x > 0 else "Brak danych"
    )
    export_df["Godziny twórcze (h)"] = df["creative_hours"]

    return export_df


# =============================================================================
# EKSPORT WORKLOGS (MIESIĄC)
# =============================================================================

def export_worklogs_to_excel(
    month_data: pd.DataFrame,
    selected_month: str,
    start_date: Any,
    end_date: Any
) -> tuple[io.BytesIO, str]:
    """
    Eksportuje worklogs miesiąca do Excel (2 arkusze: Worklogs + Podsumowanie).

    Args:
        month_data: DataFrame z danymi miesiąca
        selected_month: String miesiąca (np. "2025-01")
        start_date: Data początkowa
        end_date: Data końcowa

    Returns:
        Tuple (buffer, filename)
    """
    from helpers import calculate_creative_summary

    buffer = io.BytesIO()
    styles = ExcelStyles()

    # Przygotuj dane szczegółowe
    export_df = _prepare_excel_dataframe(month_data)

    # Przygotuj podsumowanie
    summary_df = calculate_creative_summary(month_data).reset_index()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Arkusz 1: Worklogs
        export_df.to_excel(writer, sheet_name="Worklogs", index=False)
        worksheet = writer.sheets["Worklogs"]

        set_column_widths(worksheet)
        format_worksheet_headers(worksheet, styles, len(export_df.columns))
        format_worksheet_data(
            worksheet, styles,
            row_count=len(export_df),
            column_count=len(export_df.columns)
        )

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:{chr(65 + len(export_df.columns) - 1)}{len(export_df) + 1}"

        # Arkusz 2: Podsumowanie
        summary_df.to_excel(writer, sheet_name="Podsumowanie", index=False)
        worksheet_summary = writer.sheets["Podsumowanie"]

        summary_widths = {"A": 25, "B": 18, "C": 18, "D": 18, "E": 18}
        set_column_widths(worksheet_summary, summary_widths)
        format_worksheet_headers(worksheet_summary, styles, len(summary_df.columns))

        for row_num in range(2, len(summary_df) + 2):
            for col_num in range(1, len(summary_df.columns) + 1):
                cell = worksheet_summary.cell(row=row_num, column=col_num)
                cell.font = styles.data_font
                cell.border = styles.thin_border

                if col_num == 1:
                    cell.alignment = styles.data_alignment_left
                else:
                    cell.alignment = styles.data_alignment_right
                    if col_num in [2, 3]:
                        cell.number_format = "0.00"
                    else:
                        cell.number_format = "0.0"

        worksheet_summary.freeze_panes = "A2"

    buffer.seek(0)

    # Formatuj daty do nazwy pliku
    start_str = start_date.strftime('%d') if hasattr(start_date, 'strftime') else str(start_date)
    end_str = end_date.strftime('%d') if hasattr(end_date, 'strftime') else str(end_date)
    filename = f"worklogs_{selected_month}_{start_str}-{end_str}.xlsx"

    return buffer, filename


def export_worklogs_to_csv(
    month_data: pd.DataFrame,
    selected_month: str,
    start_date: Any,
    end_date: Any
) -> tuple[str, str]:
    """
    Eksportuje worklogs miesiąca do CSV.

    Args:
        month_data: DataFrame z danymi miesiąca
        selected_month: String miesiąca
        start_date: Data początkowa
        end_date: Data końcowa

    Returns:
        Tuple (csv_data, filename)
    """
    export_df = month_data[
        ["person", "task", "key", "time_hours", "creative_percent", "creative_hours"]
    ].copy()
    export_df.columns = [
        "Osoba", "Zadanie", "Klucz", "Czas (h)", "% Twórczości", "Godziny twórcze"
    ]

    csv_data = export_df.to_csv(index=False, encoding="utf-8-sig")

    start_str = start_date.strftime('%d') if hasattr(start_date, 'strftime') else str(start_date)
    end_str = end_date.strftime('%d') if hasattr(end_date, 'strftime') else str(end_date)
    filename = f"worklogs_{selected_month}_{start_str}-{end_str}.csv"

    return csv_data, filename
