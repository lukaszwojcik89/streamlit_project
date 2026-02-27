"""
Raport Czasu Pracy i Pracy Twórczej - Streamlit Dashboard

Przetwarza raporty z Jiry (hierarchiczna struktura Level 0/1/2) i worklogs,
oblicza Creative Score oraz eksportuje dane.
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go

from helpers import (
    parse_time_to_hours,
    hours_to_hm_format,
    extract_creative_percentage,
    fix_polish_encoding,
    apply_encoding_fix_to_dataframe,
    get_top_task_per_person,
    format_display_table,
    calculate_creative_summary,
    get_dynamic_creative_filter_options,
    validate_data_structure,
    generate_executive_summary,
)
from export_utils import (
    export_to_csv,
    export_to_excel,
    export_worklogs_to_csv,
    export_worklogs_to_excel,
)
from config import (
    MAX_FILE_SIZE_MB,
    LARGE_FILE_WARNING_MB,
    TABLE_HEADERS_WITH_EMOJI,
    TABLE_HEADERS_PLAIN,
    DAY_NAMES_PL,
    DAY_ORDER,
    CHART_MIN_HEIGHT,
    CHART_ROW_HEIGHT,
)

# =============================================================================
# KONFIGURACJA STREAMLIT
# =============================================================================

st.set_page_config(
    page_title="Raport Czasu Pracy",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)


# =============================================================================
# PRZETWARZANIE DANYCH (CACHED)
# =============================================================================


@st.cache_data(show_spinner=False)
def process_excel_data(df: pd.DataFrame) -> pd.DataFrame:
    """Przetwarza dane z Excel do struktury raportu (Level 0/1/2)."""
    report_data = []
    current_user = None
    current_task = None

    for idx, row in df.iterrows():
        level = row["Level"]
        description = row["Users / Issues / Procent pracy twórczej"]
        key = row.get("Key", "")
        time_spent = row.get("Total Time Spent", "0:00")

        if level == 0:  # Użytkownik
            current_user = description
        elif level == 1 and current_user:  # Zadanie
            current_task = {
                "person": current_user,
                "task": description,
                "key": key if pd.notna(key) else "",
                "time_spent": time_spent,
                "time_hours": parse_time_to_hours(time_spent),
                "creative_percent": None,
                "creative_hours": 0.0,
            }
            report_data.append(current_task)
        elif level == 2 and current_task is not None and pd.notna(description):
            creative_percent = extract_creative_percentage(description)
            if creative_percent is not None:
                current_task["creative_percent"] = creative_percent
                current_task["creative_hours"] = (
                    creative_percent / 100
                ) * current_task["time_hours"]

    df_result = pd.DataFrame(report_data)

    if not df_result.empty:
        df_result["time_hours"] = df_result["time_hours"].astype(float)
        df_result["creative_hours"] = df_result["creative_hours"].astype(float)
        df_result["creative_percent"] = pd.to_numeric(
            df_result["creative_percent"], errors="coerce"
        )

    return df_result


@st.cache_data(show_spinner=False)
def process_worklogs_data(df: pd.DataFrame) -> pd.DataFrame:
    """Przetwarza dane z worklogs (płaski format z datami)."""
    df_work = df.copy()

    df_work["Start Date"] = pd.to_datetime(df_work["Start Date"], errors="coerce")
    df_work["time_hours"] = df_work["Time Spent"].apply(parse_time_to_hours)
    df_work["creative_percent"] = df_work["Procent pracy twórczej"].apply(
        extract_creative_percentage
    )
    df_work["creative_hours"] = (
        df_work["creative_percent"].fillna(0) / 100 * df_work["time_hours"]
    )

    df_work["person"] = df_work["Author"]
    df_work["task"] = df_work["Issue Summary"]
    df_work["key"] = df_work["Issue Key"]
    df_work["month_str"] = df_work["Start Date"].dt.strftime("%Y-%m")

    return df_work[
        [
            "person",
            "task",
            "key",
            "time_hours",
            "creative_percent",
            "creative_hours",
            "Start Date",
            "month_str",
        ]
    ]


def main():
    st.title("📊 Raport Czasu Pracy i Pracy Twórczej")
    st.markdown("---")

    # Inicjalizacja session state
    if "uploaded_files_history" not in st.session_state:
        st.session_state.uploaded_files_history = []
    if "current_file_data" not in st.session_state:
        st.session_state.current_file_data = None

    # Sidebar dla uploadowania pliku
    with st.sidebar:
        st.header("📁 Wgraj pliki")

        # Przycisk do czyszczenia cache'a
        if st.button(
            "🔄 Wyczyść cache (jeśli procenty się nie ładują)", width="stretch"
        ):
            st.cache_data.clear()
            st.success("✅ Cache wyzczyszczony! Wgraj plik ponownie.")

        st.markdown("---")

        st.subheader("📊 Raport główny")
        uploaded_file = st.file_uploader(
            "Wybierz plik Excel (.xlsx) - struktura Level 0/1/2",
            type=["xlsx"],
            key="main_report",
            help="Raport: Użytkownik (0) / Zadanie (1) / % Twórczości (2)",
        )

        st.markdown("---")

        st.subheader("📋 Worklogs (opcjonalnie)")
        worklogs_file = st.file_uploader(
            "Wgraj worklogs z datami",
            type=["xlsx"],
            key="worklogs_file",
            help="Worklogs: Start Date, Issue Key, Time Spent, Procent pracy twórczej",
        )

        st.markdown("---")

        # Walidacja rozmiaru pliku
        if uploaded_file:
            file_size_mb = uploaded_file.size / (1024 * 1024)
            if file_size_mb > 50:
                st.error(
                    f"❌ Plik zbyt duży: {file_size_mb:.1f}MB. Maksymalny rozmiar to 50MB."
                )
                uploaded_file = None
            elif file_size_mb > 10:
                st.warning(
                    f"⚠️ Duży plik: {file_size_mb:.1f}MB. Przetwarzanie może chwilę potrwać."
                )

        st.markdown("---")
        st.header("ℹ️ Informacje")
        st.markdown(
            """
        **Struktura pliku głównego (Level 0/1/2):**
        - Poziom 0: Nazwiska użytkowników
        - Poziom 1: Zadania z kluczami i czasem
        - Poziom 2: Procent pracy twórczej
        
        **Kalkulacje:**
        - Procent 90% + 10h = 9h pracy twórczej
        
        **Worklogs (opcjonalnie):**
        - Płaskie dane z datami
        - Analizy per miesiąc, timeline, export
        - Deduplikacja wspólnych Issue Key
        
        **Format procentów:** `90`, `90%`, `80.5%`
        """
        )

    if uploaded_file is not None:
        try:
            # Wczytaj dane - spróbuj różne kodowania
            with st.spinner("📂 Wczytuję plik Excel..."):
                try:
                    df_raw = pd.read_excel(uploaded_file, engine="openpyxl")
                except UnicodeDecodeError:
                    # Jeśli problem z kodowaniem, spróbuj ponownie
                    uploaded_file.seek(0)
                    df_raw = pd.read_excel(
                        uploaded_file, engine="openpyxl", encoding="utf-8"
                    )

            # Funkcja do naprawiania kodowania polskich znaków
            def fix_encoding(text):
                if pd.isna(text) or not isinstance(text, str):
                    return text

                # Mapa najczęstszych błędów kodowania
                fixes = {
                    "Ä…": "ą",
                    "Ä": "Ą",
                    "Ä‡": "ć",
                    "Ć": "Ć",
                    "Ĺ‚": "ł",
                    "Ĺ": "Ł",
                    "Ĺ›": "ś",
                    "Ĺš": "Ś",
                    "ĹĽ": "ż",
                    "Ĺ»": "Ż",
                    "Ĺş": "ź",
                    "Ĺą": "Ź",
                    "Ă³": "ó",
                    'Ă"': "Ó",
                    "Ĺ„": "ń",
                    "Ĺƒ": "Ń",
                    "Ä™": "ę",
                    "Äś": "Ę",
                    "moĹĽliwoĹ›Ä‡": "możliwość",
                    "dodaÄ‡": "dodać",
                    "hiperĹ‚Ä…cze": "hiperłącze",
                    "czcionki/tĹ‚a": "czcionki/tła",
                }

                result = str(text)
                for wrong, correct in fixes.items():
                    result = result.replace(wrong, correct)
                return result

            # Zastosuj naprawę kodowania do wszystkich kolumn tekstowych
            for col in df_raw.columns:
                if df_raw[col].dtype == "object":
                    df_raw[col] = df_raw[col].apply(fix_encoding)

            # Sprawdź strukturę
            required_columns = ["Level", "Users / Issues / Procent pracy twórczej"]
            if not all(col in df_raw.columns for col in required_columns):
                st.error(f"❌ Plik nie zawiera wymaganych kolumn: {required_columns}")
                st.info(f"Znalezione kolumny: {list(df_raw.columns)}")
                return

            # Waliduj strukturę danych
            issues, warnings = validate_data_structure(df_raw)

            # Wyświetl krytyczne błędy
            if issues:
                st.error("❌ Wykryto krytyczne problemy z plikiem:")
                for issue in issues:
                    st.error(f"  • {issue}")
                st.info("Sprawdź czy struktura pliku jest poprawna (Level 0/1/2)")
                return

            # Wyświetl ostrzeżenia
            if warnings:
                with st.expander(
                    "⚠️ Wykryto potencjalne problemy (kliknij aby zobaczyć)",
                    expanded=False,
                ):
                    for warning in warnings:
                        st.warning(f"  • {warning}")
                    st.info("Możesz kontynuować, ale wyniki mogą być niepełne.")

            # Upewnij się, że Level jest liczbą całkowitą
            df_raw["Level"] = (
                pd.to_numeric(df_raw["Level"], errors="coerce").fillna(0).astype(int)
            )

            # Debug: Pokaż przykładowe dane
            # with st.expander("🔍 Debug - Podgląd surowych danych (pierwsze 20 wierszy)"):
            #     st.dataframe(df_raw.head(20))
            #     st.write("**Kolumny:**", list(df_raw.columns))
            #     st.write("**Typy danych:**")
            #     st.write(df_raw.dtypes)

            # Przetwórz dane
            with st.spinner("⚙️ Przetwarzam dane..."):
                df_processed = process_excel_data(df_raw)

            if df_processed.empty:
                st.warning("⚠️ Nie udało się przetworzyć danych z pliku.")
                st.info("Sprawdź czy struktura pliku jest poprawna (Level 0/1/2)")
                return

            # Wczytaj i przetwórz worklogs jeśli został wgrany
            df_worklogs_by_month = None
            months_available = []

            if worklogs_file is not None:
                try:
                    with st.spinner("📋 Przetwarzam worklogs..."):
                        df_worklogs_raw = pd.read_excel(
                            worklogs_file, engine="openpyxl"
                        )
                        df_worklogs = process_worklogs_data(df_worklogs_raw)

                        if not df_worklogs.empty:
                            # Info o wspólnych kluczach (bez usuwania - worklogs
                            # to szczegóły z datami, nie duplikaty)
                            worklogs_keys = set(df_worklogs["key"].dropna().unique())
                            report_keys = set(df_processed["key"].dropna().unique())
                            duplicated_keys = worklogs_keys & report_keys

                            if duplicated_keys:
                                st.info(
                                    f"ℹ️ {len(duplicated_keys)} wspólnych Issue Keys "
                                    f"między raportem a worklogs (to normalne - "
                                    f"worklogs zawiera szczegóły z datami)"
                                )

                            # Agreguj po miesiącach
                            df_worklogs_by_month = {
                                month: group.copy()
                                for month, group in df_worklogs.groupby("month_str")
                            }
                            months_available = sorted(
                                df_worklogs_by_month.keys(), reverse=True
                            )
                            st.success(
                                f"✅ Worklogs przetworzony! "
                                f"{len(df_worklogs)} wpisów, "
                                f"miesiące: {', '.join(months_available)}"
                            )
                except Exception as e:
                    st.error(f"❌ Nie udało się przetworzyć worklogs: {str(e)}")
                    with st.expander("🐞 Szczegóły błędu"):
                        import traceback
                        st.code(traceback.format_exc())

            # Debug: Pokaż przetworzone dane
            # with st.expander("🔍 Debug - Przetworzone dane (pierwsze 10 wierszy)"):
            #     st.dataframe(df_processed.head(10))
            #     st.write(f"**Liczba wierszy:** {len(df_processed)}")
            #     st.write(f"**Kolumny:** {list(df_processed.columns)}")
            #     st.write(f"**Zadania z % twórczości:** {df_processed['creative_percent'].notna().sum()}")
            #     st.write(f"**Zadania bez % twórczości:** {df_processed['creative_percent'].isna().sum()}")
            #
            #     # Debug Level 2 data
            #     st.write("**Debug: Dane Level 2 (procenty pracy twórczej) z surowego pliku:**")
            #     level2_data = df_raw[df_raw['Level'] == 2][['Users / Issues / Procent pracy twórczej', 'Key']].head(20)
            #     st.dataframe(level2_data, use_container_width=True)
            #
            #     # Debug ekstrakcji procentów
            #     st.write("**Debug: Ekstrahowane procenty:**")
            #     level2_texts = df_raw[df_raw['Level'] == 2]['Users / Issues / Procent pracy twórczej'].dropna().head(20)
            #     for idx, text in enumerate(level2_texts):
            #         extracted = extract_creative_percentage(text)
            #         st.write(f"`{text}` → {extracted}")

            # Statystyki główne
            col1, col2, col3, col4 = st.columns(4)

            with col1:
                st.metric("👥 Liczba osób", df_processed["person"].nunique())

            with col2:
                st.metric("📋 Liczba zadań", len(df_processed))

            with col3:
                total_hours = df_processed["time_hours"].sum()
                st.metric("⏰ Łączne godziny", f"{total_hours:.1f}h")

            with col4:
                creative_tasks = df_processed["creative_percent"].notna().sum()
                st.metric("🎨 Zadania z % twórczości", creative_tasks)

            st.markdown("---")

            # ===========================================================================
            # TABS: GŁÓWNY INTERFEJS
            # ===========================================================================
            tab_report, tab_worklogs, tab_help = st.tabs([
                "📊 Raport główny",
                "📋 Worklogs (Sprint)",
                "❓ Pomoc"
            ])

            # ===========================================================================
            # TAB 1: RAPORT GŁÓWNY
            # ===========================================================================
            with tab_report:

                # === 🎯 TOP ZADANIA PER OSOBA (WIDOCZNE - KLUCZOWA INFORMACJA!) ===
                st.markdown("## 🎯 Top Zadanie Każdego Użytkownika")
                st.caption("Score = godziny twórcze × (% twórczości / 100) — nagradza kombinację długiego zaangażowania i WYSOKIEJ kreatywności")

                # Znajdź top zadanie dla każdego użytkownika
                most_creative_tasks = []
                for person in sorted(df_processed["person"].unique()):
                    person_data = df_processed[df_processed["person"] == person]

                    # Oblicz score: creative_hours * (creative_percent / 100)
                    creative_data = person_data[person_data["creative_hours"] > 0].copy()

                    if not creative_data.empty:
                        creative_data["score"] = (
                            creative_data["creative_hours"]
                            * creative_data["creative_percent"] / 100
                        )
                        best_task = creative_data.nlargest(1, "score").iloc[0]
                        most_creative_tasks.append(
                            {
                                "person": best_task["person"],
                                "task": best_task["task"],
                                "key": best_task["key"],
                                "time_hours": best_task["time_hours"],
                                "creative_percent": best_task["creative_percent"],
                                "creative_hours": best_task["creative_hours"],
                                "score": best_task["score"],
                                "has_creative_data": True,
                            }
                        )
                    else:
                        # Brak danych o twórczości - bierz najdłuższe zadanie
                        best_task = person_data.nlargest(1, "time_hours").iloc[0]
                        most_creative_tasks.append(
                            {
                                "person": best_task["person"],
                                "task": best_task["task"],
                                "key": best_task["key"],
                                "time_hours": best_task["time_hours"],
                                "creative_percent": best_task["creative_percent"],
                                "creative_hours": best_task["creative_hours"],
                                "score": 0.0,
                                "has_creative_data": False,
                            }
                        )

                if most_creative_tasks:
                    most_creative_df = pd.DataFrame(most_creative_tasks)

                    # Posortuj po score
                    most_creative_df_sorted = most_creative_df.sort_values(
                        by="score", ascending=False
                    )

                    most_creative_display = most_creative_df_sorted.copy()
                    most_creative_display["time_hours"] = most_creative_display[
                        "time_hours"
                    ].apply(lambda x: f"{x:.1f}h")
                    most_creative_display["creative_percent"] = most_creative_display[
                        "creative_percent"
                    ].apply(lambda x: f"{int(x)}%" if pd.notna(x) else "—")
                    most_creative_display["creative_hours"] = most_creative_display[
                        "creative_hours"
                    ].apply(lambda x: f"{x:.1f}h" if x > 0 else "—")
                    most_creative_display["score"] = most_creative_display[
                        "score"
                    ].apply(lambda x: f"{x:.2f}" if x > 0 else "—")
                    most_creative_display["status"] = most_creative_display[
                        "has_creative_data"
                    ].apply(
                        lambda x: "✨ Twórcze" if x else "⏰ Brak danych (najdłuższe)"
                    )

                    display_cols = [
                        "person",
                        "task",
                        "key",
                        "time_hours",
                        "creative_percent",
                        "creative_hours",
                        "score",
                        "status",
                    ]
                    display_names = [
                        "👤 Osoba",
                        "📋 Zadanie",
                        "🔑 Klucz",
                        "⏰ Czas",
                        "🎨 %",
                        "✨ Godz. twórcze",
                        "🏆 Score",
                        "📊 Typ",
                    ]

                    st.dataframe(
                        most_creative_display[display_cols].rename(
                            columns=dict(zip(display_cols, display_names))
                        ),
                        hide_index=True,
                        width="stretch",
                    )

                    # Wykres
                    fig_most_creative = px.bar(
                        most_creative_df_sorted,
                        x="score",
                        y="person",
                        orientation="h",
                        title="Creative Score — balans czasu i kreatywności",
                        labels={"score": "Score", "person": "Osoba"},
                        color="creative_percent",
                        color_continuous_scale="Viridis",
                        hover_data=["time_hours", "creative_hours", "creative_percent"],
                    )
                    fig_most_creative.update_layout(
                        height=max(300, len(most_creative_tasks) * 40),
                        xaxis_title="Creative Score",
                        yaxis_title="",
                        coloraxis_colorbar_title="% Twórczości",
                    )
                    st.plotly_chart(fig_most_creative, width="stretch")

                # === SEKCJA 2: DODATKOWE INFORMACJE (EXPANDER) ===
                with st.expander("📊 Dodatkowe informacje", expanded=False):
                    # Top 5 najdłuższych zadań
                    st.markdown("**🔝 Top 5 najdłuższych zadań**")
                    top_tasks = df_processed.nlargest(5, "time_hours")[
                        ["person", "task", "key", "time_hours", "creative_percent"]
                    ]
                    top_tasks_display = top_tasks.copy()
                    top_tasks_display["time_hours"] = top_tasks_display["time_hours"].apply(
                        lambda x: f"{x:.1f}h"
                    )
                    top_tasks_display["creative_percent"] = top_tasks_display[
                        "creative_percent"
                    ].apply(lambda x: f"{int(x)}%" if pd.notna(x) else "-")
                    top_tasks_display.columns = [
                        "👤 Osoba",
                        "📋 Zadanie",
                        "🔑 Klucz",
                        "⏰ Czas",
                        "🎨 % Twórczości",
                    ]
                    st.dataframe(
                        top_tasks_display, hide_index=True, width="stretch"
                    )
                    
                    st.markdown("---")
                    
                    # Statystyki
                    st.markdown("**📈 Statystyki ogólne**")
                    col1, col2, col3 = st.columns(3)

                    with col1:
                        # Średni czas na zadanie
                        avg_time = df_processed["time_hours"].mean()
                        st.metric("⏱️ Średni czas na zadanie", f"{avg_time:.1f}h")

                        # Mediana czasu
                        median_time = df_processed["time_hours"].median()
                        st.metric("📊 Mediana czasu", f"{median_time:.1f}h")

                    with col2:
                        # Średni procent twórczości
                        avg_creative = df_processed["creative_percent"].mean()
                        if pd.notna(avg_creative):
                            st.metric("🎨 Średni % twórczości", f"{avg_creative:.0f}%")
                        else:
                            st.metric("🎨 Średni % twórczości", "Brak danych")

                        # Pokrycie danymi
                        coverage = (
                            df_processed["creative_percent"].notna().sum()
                            / len(df_processed)
                            * 100
                        )
                        st.metric("📋 Pokrycie danymi", f"{coverage:.0f}%")

                    with col3:
                        # Najbardziej aktywna osoba
                        most_active = (
                            df_processed.groupby("person")["time_hours"].sum().idxmax()
                        )
                        most_active_hours = (
                            df_processed.groupby("person")["time_hours"].sum().max()
                        )
                        st.metric("🏆 Najbardziej aktywny", most_active)
                        st.caption(f"{most_active_hours:.1f}h łącznie")

            # === SEKCJA 3: SZCZEGÓŁOWE DANE (EXPANDER) ===
            with st.expander("🔍 Szczegółowe dane", expanded=False):
                # Filtry
                col1, col2, col3 = st.columns(3)

                with col1:
                    selected_person = st.selectbox(
                        "👤 Wybierz osobę:",
                        ["Wszystkie"] + sorted(df_processed["person"].unique().tolist()),
                    )

                with col2:
                    creative_filter = st.selectbox(
                        "🎨 Filtruj po pracy twórczej:",
                        [
                            "Wszystkie",
                            "Z danymi",
                            "Bez danych",
                            "100%",
                            "90%",
                            "80%",
                            "60%",
                            "50%",
                            "10%",
                            "0%",
                        ],
                    )

                with col3:
                    search_term = st.text_input("🔍 Szukaj w zadaniach:", "")

                # Filtrowanie danych
                df_filtered = df_processed.copy()

                if selected_person != "Wszystkie":
                    df_filtered = df_filtered[df_filtered["person"] == selected_person]

                if creative_filter != "Wszystkie":
                    if creative_filter == "Z danymi":
                        df_filtered = df_filtered[df_filtered["creative_percent"].notna()]
                    elif creative_filter == "Bez danych":
                        df_filtered = df_filtered[df_filtered["creative_percent"].isna()]
                    else:
                        percent_val = int(creative_filter.replace("%", ""))
                        df_filtered = df_filtered[
                            df_filtered["creative_percent"] == percent_val
                        ]

                if search_term:
                    df_filtered = df_filtered[
                        df_filtered["task"].str.contains(search_term, case=False, na=False)
                        | df_filtered["key"].str.contains(search_term, case=False, na=False)
                    ]

                # Wyświetl tabelę
                st.markdown("**📋 Tabela danych**")

                # Przygotuj dane do wyświetlenia
                display_df = df_filtered.copy()

                # Konwertuj typy danych aby uniknąć błędów pyarrow
                display_df["time_hours"] = display_df["time_hours"].astype(float)
                display_df["creative_hours"] = display_df["creative_hours"].astype(float)
                # Konwertuj creative_percent na float, None zamień na NaN
                display_df["creative_percent"] = pd.to_numeric(
                    display_df["creative_percent"], errors="coerce"
                )

                display_df["creative_percent_display"] = display_df[
                    "creative_percent"
                ].apply(lambda x: f"{int(x)}%" if pd.notna(x) else "Brak danych")
                display_df["creative_hours_display"] = display_df["creative_hours"].apply(
                    lambda x: f"{x:.1f}h" if pd.notna(x) else "Brak danych"
                )
                display_df["time_display"] = display_df["time_hours"].apply(
                    lambda x: f"{x:.1f}h"
                )

                # Kolumny do wyświetlenia
                columns_to_show = [
                    "person",
                    "task",
                    "key",
                    "time_display",
                    "creative_percent_display",
                    "creative_hours_display",
                ]

                # Nazwy kolumn bez emotek do eksportu
                export_column_names = [
                    "Osoba",
                    "Zadanie",
                    "Klucz",
                    "Czas",
                    "Procent twórczości",
                    "Godziny twórcze",
                ]

                st.dataframe(
                    display_df[columns_to_show],
                    column_config={
                        "person": st.column_config.TextColumn("👤 Osoba", width="medium"),
                        "task": st.column_config.TextColumn("📋 Zadanie", width="large"),
                        "key": st.column_config.TextColumn("🔑 Klucz", width="small"),
                        "time_display": st.column_config.TextColumn(
                            "⏰ Czas", width="small"
                        ),
                        "creative_percent_display": st.column_config.TextColumn(
                            "🎨 % Twórczości", width="small"
                        ),
                        "creative_hours_display": st.column_config.TextColumn(
                            "🎯 Godziny twórcze", width="small"
                        ),
                    },
                    width="stretch",
                    hide_index=True,
                )

                # Wykresy
                st.markdown("---")
                st.markdown("**📊 Wykresy analityczne**")

                col1, col2 = st.columns(2)

                with col1:
                    st.markdown("**Czas pracy na osobę**")
                    person_hours = (
                        df_filtered.groupby("person")["time_hours"]
                        .sum()
                        .sort_values(ascending=True)
                    )

                    fig1 = px.bar(
                        x=person_hours.values,
                        y=person_hours.index,
                        orientation="h",
                        title="Łączne godziny pracy",
                        labels={"x": "Godziny", "y": "Osoba"},
                    )
                    fig1.update_layout(height=400)
                    st.plotly_chart(
                        fig1,
                        width="stretch",
                        config={"displayModeBar": True},
                    )

                with col2:
                    st.markdown("**Rozkład pracy twórczej**")
                    creative_data = df_filtered.dropna(subset=["creative_percent"])
                    if not creative_data.empty:
                        creative_counts = (
                            creative_data["creative_percent"].value_counts().sort_index()
                        )

                        fig2 = px.pie(
                            values=creative_counts.values,
                            names=[f"{x}%" for x in creative_counts.index],
                            title="Procent zadań według poziomu twórczości",
                        )
                        fig2.update_layout(height=400)
                        st.plotly_chart(
                            fig2,
                            width="stretch",
                            config={"displayModeBar": True},
                        )
                    else:
                        st.info("Brak danych o pracy twórczej do wyświetlenia.")

                # Podsumowanie pracy twórczej
                st.markdown("**🎯 Podsumowanie pracy twórczej**")

                creative_summary = (
                    df_filtered.groupby("person")
                    .agg(
                        {
                            "time_hours": "sum",
                            "creative_hours": "sum",
                            "creative_percent": lambda x: x.dropna().count(),  # liczba zadań z danymi
                        }
                    )
                    .round(2)
                )

                # Oblicz czas TYLKO dla zadań z danymi o twórczości
                time_hours_with_data = df_filtered[df_filtered["creative_percent"].notna()].groupby("person")["time_hours"].sum()
                
                # Główny wskaźnik - % twórczości ze ZGRUPOWANYCH GODZIN (gdzie mamy dane)
                creative_summary["creative_ratio"] = (
                    creative_summary["creative_hours"]
                    / time_hours_with_data
                    * 100
                ).round(1)

                # Wskaźnik pokrycia
                total_tasks = df_filtered.groupby("person").size()
                creative_summary["coverage"] = (
                    creative_summary["creative_percent"] / total_tasks * 100
                ).round(0)

                # Wybierz kolumny
                creative_summary = creative_summary[
                    ["time_hours", "creative_hours", "creative_ratio", "coverage"]
                ]
                creative_summary.columns = [
                    "Łączne godziny",
                    "Godziny twórcze",
                    "% Pracy twórczej",
                    "Pokrycie danymi",
                ]

                st.dataframe(
                    creative_summary,
                    column_config={
                        "Łączne godziny": st.column_config.NumberColumn(
                            "Łączne godziny",
                            format="%.1f h",
                            help="Całkowity czas pracy danej osoby",
                        ),
                        "Godziny twórcze": st.column_config.NumberColumn(
                            "Godziny twórcze",
                            format="%.1f h",
                            help="Faktyczne godziny pracy twórczej (procent × czas zadania)",
                        ),
                        "% Pracy twórczej": st.column_config.NumberColumn(
                            "% Pracy twórczej",
                            format="%.1f%%",
                            help="Rzeczywisty procent: godziny twórcze / łączne godziny",
                        ),
                        "Pokrycie danymi": st.column_config.NumberColumn(
                            "Pokrycie danymi",
                            format="%.0f%%",
                            help="Procent zadań z przypisanymi danymi o twórczości",
                        ),
                    },
                    width="stretch",
                )

                # Dodatkowe wykresy
                st.markdown("---")
                st.markdown("**📊 Więcej analiz**")

                col1, col2 = st.columns(2)

                with col1:
                    # Heatmapa: Osoba vs Typ pracy twórczej
                    st.markdown("**Rozkład typów pracy twórczej per osoba**")
                    creative_data = df_filtered.dropna(subset=["creative_percent"])
                    if not creative_data.empty and len(creative_data) > 0:
                        # Utwórz pivot table
                        heatmap_data = (
                            creative_data.groupby(["person", "creative_percent"])
                            .size()
                            .reset_index(name="count")
                        )
                        heatmap_pivot = heatmap_data.pivot(
                            index="person", columns="creative_percent", values="count"
                        ).fillna(0)

                        fig_heatmap = px.imshow(
                            heatmap_pivot,
                            labels=dict(x="% Twórczości", y="Osoba", color="Liczba zadań"),
                            x=[f"{int(x)}%" for x in heatmap_pivot.columns],
                            y=heatmap_pivot.index,
                            color_continuous_scale="Blues",
                            aspect="auto",
                        )
                        fig_heatmap.update_layout(height=400)
                        st.plotly_chart(fig_heatmap, width="stretch")
                    else:
                        st.info("Brak danych do wyświetlenia heatmapy")

                with col2:
                    # Porównanie czasu pracy vs czasu twórczego
                    st.markdown("**Czas pracy vs Czas twórczy (per osoba)**")
                    comparison_data = (
                        df_filtered.groupby("person")
                        .agg({"time_hours": "sum", "creative_hours": "sum"})
                        .reset_index()
                    )
                    comparison_data.columns = ["Osoba", "Łączny czas", "Czas twórczy"]

                    fig_comparison = go.Figure()
                    fig_comparison.add_trace(
                        go.Bar(
                            name="Łączny czas",
                            x=comparison_data["Osoba"],
                            y=comparison_data["Łączny czas"],
                            marker_color="lightblue",
                        )
                    )
                    fig_comparison.add_trace(
                        go.Bar(
                            name="Czas twórczy",
                            x=comparison_data["Osoba"],
                            y=comparison_data["Czas twórczy"],
                            marker_color="darkblue",
                        )
                    )
                    fig_comparison.update_layout(
                        barmode="group",
                        height=400,
                        yaxis_title="Godziny",
                        xaxis_title="Osoba",
                        legend=dict(
                            orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1
                        ),
                    )
                    st.plotly_chart(fig_comparison, width="stretch")

            # === SEKCJA 5: EKSPORT DANYCH ===
            st.markdown("---")
            st.markdown("## 📥 Eksport Danych")

            col1, col2 = st.columns(2)

            with col1:
                # Eksport CSV z polskimi znakami (bez emotek)
                export_df = display_df[columns_to_show].copy()
                export_df.columns = export_column_names  # Użyj nazw bez emotek

                csv = export_df.to_csv(index=False, encoding="utf-8-sig")
                st.download_button(
                    label="📋 Pobierz jako CSV",
                    data=csv,
                    file_name=f"raport_pracy_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv",
                    width="stretch",
                )

            with col2:
                # Eksport Excel z profesjonalnym formatowaniem
                import io

                buffer = io.BytesIO()

                export_df_excel = display_df[columns_to_show].copy()
                export_df_excel.columns = export_column_names  # Użyj nazw bez emotek

                # Konwertuj kolumnę czasu na format HH:MM
                export_df_excel["Czas"] = display_df["time_hours"].apply(
                    hours_to_hm_format
                )

                # Dodaj kolumnę "Czas (h)" - czas jako liczba w godzinach
                export_df_excel.insert(
                    export_df_excel.columns.get_loc("Czas") + 1,
                    "Czas (h)",
                    display_df["time_hours"],
                )

                # Przygotuj kolumnę "Procent twórczości" jako liczba całkowita (0-100)
                export_df_excel["Procent twórczości"] = display_df["creative_percent"]

                # Dodaj kolumnę "Godziny twórcze (h)" - godziny jako liczba
                export_df_excel.insert(
                    export_df_excel.columns.get_loc("Godziny twórcze") + 1,
                    "Godziny twórcze (h)",
                    display_df["creative_hours"],
                )

                with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                    # Arkusz 1: Szczegółowe dane
                    export_df_excel.to_excel(
                        writer, sheet_name="Raport pracy", index=False
                    )

                    # Formatowanie arkusza szczegółowego
                    worksheet = writer.sheets["Raport pracy"]

                    # Dostosuj szerokość kolumn
                    column_widths = {
                        "A": 25,  # Osoba
                        "B": 50,  # Zadanie
                        "C": 15,  # Klucz
                        "D": 12,  # Czas
                        "E": 12,  # Czas (h)
                        "F": 18,  # % Twórczości
                        "G": 18,  # Godziny twórcze
                        "H": 18,  # Godziny twórcze (h)
                    }

                    for col, width in column_widths.items():
                        worksheet.column_dimensions[col].width = width

                    # Formatowanie profesjonalne
                    from openpyxl.styles import (
                        Font,
                        PatternFill,
                        Alignment,
                        Border,
                        Side,
                    )

                    # Style nagłówka
                    header_font = Font(
                        name="Calibri", size=11, bold=True, color="FFFFFF"
                    )
                    header_fill = PatternFill(
                        start_color="2F5597", end_color="2F5597", fill_type="solid"
                    )
                    header_alignment = Alignment(horizontal="center", vertical="center")

                    # Ramki
                    thin_border = Border(
                        left=Side(style="thin", color="E0E0E0"),
                        right=Side(style="thin", color="E0E0E0"),
                        top=Side(style="thin", color="E0E0E0"),
                        bottom=Side(style="thin", color="E0E0E0"),
                    )

                    # Formatuj nagłówki
                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border

                    # Style dla danych
                    data_font = Font(name="Calibri", size=10)
                    data_alignment_left = Alignment(
                        horizontal="left", vertical="center"
                    )
                    data_alignment_center = Alignment(
                        horizontal="center", vertical="center"
                    )
                    data_alignment_right = Alignment(
                        horizontal="right", vertical="center"
                    )

                    # Kolory dla procentów twórczości
                    color_red = PatternFill(
                        start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
                    )  # 0-50%
                    color_yellow = PatternFill(
                        start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"
                    )  # 50-80%
                    color_green = PatternFill(
                        start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"
                    )  # 80-100%

                    # Formatuj dane i dodaj kolorowanie
                    for row_num in range(2, len(export_df_excel) + 2):
                        # Pobierz wartość procentu twórczości
                        creative_percent_value = worksheet.cell(
                            row=row_num, column=6
                        ).value

                        for col_num in range(1, len(export_df_excel.columns) + 1):
                            cell = worksheet.cell(row=row_num, column=col_num)
                            cell.font = data_font
                            cell.border = thin_border

                            # Wyrównanie i formatowanie w zależności od kolumny
                            if col_num in [1, 2]:  # Osoba, Zadanie
                                cell.alignment = data_alignment_left
                            elif col_num in [3]:  # Klucz
                                cell.alignment = data_alignment_center
                            elif col_num == 5:  # Czas (h)
                                cell.alignment = data_alignment_right
                                cell.number_format = "0.00"
                            elif col_num == 6:  # % Twórczości
                                cell.alignment = data_alignment_right
                                cell.number_format = "0"
                                # Kolorowanie na podstawie wartości
                                if (
                                    pd.notna(creative_percent_value)
                                    and creative_percent_value != ""
                                ):
                                    try:
                                        percent = float(creative_percent_value)
                                        if percent <= 50:
                                            cell.fill = color_red
                                        elif percent <= 80:
                                            cell.fill = color_yellow
                                        else:
                                            cell.fill = color_green
                                    except (ValueError, TypeError):
                                        pass
                            elif col_num == 8:  # Godziny twórcze (h)
                                cell.alignment = data_alignment_right
                                cell.number_format = "0.00"
                            else:  # Pozostałe kolumny (Czas HH:MM, Godziny twórcze text)
                                cell.alignment = data_alignment_right

                    # Zamrożenie pierwszego wiersza
                    worksheet.freeze_panes = "A2"

                    # Filtr automatyczny
                    worksheet.auto_filter.ref = f"A1:{chr(65 + len(export_df_excel.columns) - 1)}{len(export_df_excel) + 1}"

                    # Arkusz 2: Podsumowanie
                    summary_data = creative_summary.reset_index()
                    summary_data.to_excel(
                        writer, sheet_name="Podsumowanie", index=False
                    )

                    worksheet_summary = writer.sheets["Podsumowanie"]

                    # Szerokości kolumn dla podsumowania
                    worksheet_summary.column_dimensions["A"].width = 25
                    worksheet_summary.column_dimensions["B"].width = 18
                    worksheet_summary.column_dimensions["C"].width = 18
                    worksheet_summary.column_dimensions["D"].width = 18
                    worksheet_summary.column_dimensions["E"].width = 18

                    # Formatuj nagłówki podsumowania
                    for cell in worksheet_summary[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border

                    # Formatuj dane podsumowania
                    for row_num in range(2, len(summary_data) + 2):
                        for col_num in range(1, len(summary_data.columns) + 1):
                            cell = worksheet_summary.cell(row=row_num, column=col_num)
                            cell.font = data_font
                            cell.border = thin_border

                            if col_num == 1:  # Osoba
                                cell.alignment = data_alignment_left
                            else:  # Liczby
                                cell.alignment = data_alignment_right
                                if col_num in [2, 3]:  # Godziny
                                    cell.number_format = "0.00"
                                else:  # Procenty
                                    cell.number_format = "0.0"

                    worksheet_summary.freeze_panes = "A2"

                buffer.seek(0)

                st.download_button(
                    label="📊 Pobierz jako Excel (2 arkusze: Raport + Podsumowanie)",
                    data=buffer,
                    file_name=f"raport_pracy_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    width="stretch",
                )
            
            # === TAB 2: WORKLOGS (JEŚLI DOSTĘPNE) ===
            if months_available and df_worklogs_by_month:
                st.markdown("---")
                st.markdown("## 📋 Analizy per Miesiąc (Worklogs)")

                selected_month = st.selectbox(
                    "Wybierz miesiąc do analizy",
                    months_available,
                    help="Analiza pełna miesiąca lub okresu dostępnych danych",
                )

                if selected_month in df_worklogs_by_month:
                    month_data = df_worklogs_by_month[selected_month]

                    # Oblicz range dat
                    start_date = month_data["Start Date"].min()
                    end_date = month_data["Start Date"].max()

                    # Sprawdź czy miesiąc kompletny
                    month_obj = pd.to_datetime(selected_month + "-01")
                    first_day = month_obj.replace(day=1)
                    last_day = (month_obj + pd.DateOffset(months=1)).replace(
                        day=1
                    ) - pd.Timedelta(days=1)

                    is_complete = (
                        start_date.date() <= first_day.date()
                        and end_date.date() >= last_day.date()
                    )
                    completeness_label = (
                        "✅ Pełny miesiąc"
                        if is_complete
                        else f"⚠️ Część miesiąca ({start_date.date()} do {end_date.date()})"
                    )

                    # Nagłówek z datami
                    col1, col2, col3 = st.columns([2, 2, 1])
                    with col1:
                        st.write(
                            f"**Okres:** {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
                        )
                    with col2:
                        st.write(f"**Status:** {completeness_label}")
                    with col3:
                        st.write(f"**Dni:** {(end_date - start_date).days + 1}")

                    st.markdown("---")

                    # Statystyki miesiąca
                    with st.expander("📈 Statystyki miesiąca", expanded=True):
                        stat_col1, stat_col2, stat_col3, stat_col4 = st.columns(4)

                        with stat_col1:
                            total_hours = month_data["time_hours"].sum()
                            st.metric("⏰ Łączne godziny", f"{total_hours:.1f}h")

                        with stat_col2:
                            working_days = month_data["Start Date"].dt.date.nunique()
                            avg_per_day = (
                                total_hours / working_days if working_days > 0 else 0
                            )
                            st.metric(
                                "📅 Średnio/dzień",
                                f"{avg_per_day:.1f}h",
                                delta=f"{working_days} dni pracy",
                            )

                        with stat_col3:
                            creative_hours = month_data["creative_hours"].sum()
                            avg_creative_pct = (
                                (creative_hours / total_hours * 100)
                                if total_hours > 0
                                else 0
                            )
                            st.metric(
                                "🎨 Średni % twórczości", f"{avg_creative_pct:.0f}%"
                            )

                        with stat_col4:
                            people_count = month_data["person"].nunique()
                            st.metric("👥 Liczba osób", people_count)

                        # Dodatkowe statystyki
                        col1, col2 = st.columns(2)

                        with col1:
                            st.metric(
                                "✨ Łączne godziny twórcze", f"{creative_hours:.1f}h"
                            )

                        with col2:
                            covered_tasks = month_data["creative_percent"].notna().sum()
                            total_tasks = len(month_data)
                            coverage = (
                                (covered_tasks / total_tasks * 100)
                                if total_tasks > 0
                                else 0
                            )
                            st.metric(
                                "📊 Pokrycie danymi",
                                f"{coverage:.0f}%",
                                delta=f"{covered_tasks}/{total_tasks} zadań",
                            )

                        # Top 5 osób
                        st.markdown("**🏆 Top 5 osób (po godzinach)**")
                        top_people = (
                            month_data.groupby("person")["time_hours"].sum().nlargest(5)
                        )
                        for i, (name, hours) in enumerate(top_people.items(), 1):
                            st.write(f"{i}. **{name}** - {hours:.1f}h")

                        # Top 5 zadań
                        st.markdown("**🎯 Top 5 zadań (po godzinach)**")
                        top_tasks = month_data.nlargest(5, "time_hours")[
                            ["task", "time_hours", "creative_percent"]
                        ]
                        for i, (idx, row) in enumerate(top_tasks.iterrows(), 1):
                            creative_str = (
                                f"{int(row['creative_percent'])}%"
                                if pd.notna(row["creative_percent"])
                                else "—"
                            )
                            st.write(
                                f"{i}. **{row['task'][:60]}...** - {row['time_hours']:.1f}h ({creative_str})"
                            )

                    # Timeline - wykres
                    st.markdown("---")
                    st.subheader("📊 Timeline - Godziny pracy per dzień i osoba")

                    # Agreguj po dniu i osobie
                    timeline_data = month_data.copy()
                    timeline_data["date"] = timeline_data["Start Date"].dt.date
                    daily_person = (
                        timeline_data.groupby(["date", "person"])["time_hours"]
                        .sum()
                        .reset_index()
                    )
                    daily_person = daily_person.sort_values("date")

                    # Stacked bar chart
                    fig_timeline = px.bar(
                        daily_person,
                        x="date",
                        y="time_hours",
                        color="person",
                        title=f"Rozkład godzin pracy dzień po dniu - {selected_month}",
                        labels={
                            "time_hours": "Godziny",
                            "date": "Data",
                            "person": "Osoba",
                        },
                        barmode="stack",
                    )
                    fig_timeline.update_layout(
                        height=400,
                        xaxis_title="Data",
                        yaxis_title="Godziny",
                        hovermode="x unified",
                    )
                    st.plotly_chart(fig_timeline, width="stretch")

                    # Najbardziej kreatywne zadania per osoba w danym miesiącu
                    st.markdown("---")
                    st.subheader("� Top zadanie per osoba (sprint)")
                    st.caption(
                        f"Dla każdej osoby w {selected_month}: zadanie z najlepszym "
                        f"balansem czasu i kreatywności. "
                        f"Score = godz. twórcze × (% / 100)."
                    )

                    most_creative_month = []
                    for person in sorted(month_data["person"].unique()):
                        person_data = month_data[month_data["person"] == person]
                        creative_data_p = person_data[person_data["creative_hours"] > 0].copy()

                        if not creative_data_p.empty:
                            creative_data_p["score"] = (
                                creative_data_p["creative_hours"]
                                * creative_data_p["creative_percent"] / 100
                            )
                            best = creative_data_p.nlargest(1, "score").iloc[0]
                            most_creative_month.append({
                                "person": best["person"],
                                "task": best["task"],
                                "key": best["key"],
                                "time_hours": best["time_hours"],
                                "creative_percent": best["creative_percent"],
                                "creative_hours": best["creative_hours"],
                                "score": best["score"],
                                "has_creative_data": True,
                            })
                        else:
                            best = person_data.nlargest(1, "time_hours").iloc[0]
                            most_creative_month.append({
                                "person": best["person"],
                                "task": best["task"],
                                "key": best["key"],
                                "time_hours": best["time_hours"],
                                "creative_percent": best["creative_percent"],
                                "creative_hours": best.get("creative_hours", 0),
                                "score": 0.0,
                                "has_creative_data": False,
                            })

                    if most_creative_month:
                        mc_df = pd.DataFrame(most_creative_month)
                        mc_df_sorted = mc_df.sort_values(
                            by="score", ascending=False
                        )

                        mc_display = mc_df_sorted.copy()
                        mc_display["time_hours"] = mc_display["time_hours"].apply(lambda x: f"{x:.1f}h")
                        mc_display["creative_percent"] = mc_display["creative_percent"].apply(
                            lambda x: f"{int(x)}%" if pd.notna(x) else "—"
                        )
                        mc_display["creative_hours"] = mc_display["creative_hours"].apply(
                            lambda x: f"{x:.1f}h" if x > 0 else "—"
                        )
                        mc_display["score"] = mc_display["score"].apply(
                            lambda x: f"{x:.2f}" if x > 0 else "—"
                        )
                        mc_display["status"] = mc_display["has_creative_data"].apply(
                            lambda x: "✨ Twórcze" if x else "⏰ Najdłuższe"
                        )

                        st.dataframe(
                            mc_display[["person", "task", "key", "time_hours", "creative_percent", "creative_hours", "score", "status"]].rename(
                                columns={
                                    "person": "👤 Osoba",
                                    "task": "📋 Zadanie",
                                    "key": "🔑 Klucz",
                                    "time_hours": "⏰ Czas",
                                    "creative_percent": "🎨 %",
                                    "creative_hours": "✨ Godz. twórcze",
                                    "score": "🏆 Score",
                                    "status": "📊 Typ",
                                }
                            ),
                            hide_index=True,
                            width="stretch",
                        )

                        # Wykres
                        fig_mc = px.bar(
                            mc_df_sorted,
                            x="score",
                            y="person",
                            orientation="h",
                            title=f"Creative Score — top zadanie per osoba ({selected_month})",
                            labels={"score": "Score", "person": "Osoba"},
                            color="creative_percent",
                            color_continuous_scale="Viridis",
                            hover_data=["time_hours", "creative_hours", "creative_percent"],
                        )
                        fig_mc.update_layout(
                            height=max(300, len(most_creative_month) * 40),
                            xaxis_title="Creative Score",
                            yaxis_title="",
                            coloraxis_colorbar_title="% Twórczości",
                        )
                        st.plotly_chart(fig_mc, width="stretch")

                    # Rozkład po dniach tygodnia
                    st.markdown("---")
                    st.subheader("📅 Rozkład po dniach tygodnia")

                    timeline_data["day_name"] = timeline_data[
                        "Start Date"
                    ].dt.day_name()
                    day_order = [
                        "Monday",
                        "Tuesday",
                        "Wednesday",
                        "Thursday",
                        "Friday",
                        "Saturday",
                        "Sunday",
                    ]
                    day_names_pl = {
                        "Monday": "Poniedziałek",
                        "Tuesday": "Wtorek",
                        "Wednesday": "Środa",
                        "Thursday": "Czwartek",
                        "Friday": "Piątek",
                        "Saturday": "Sobota",
                        "Sunday": "Niedziela",
                    }

                    daily_weekday = timeline_data.groupby("day_name")["time_hours"].agg(
                        ["sum", "mean", "count"]
                    )
                    daily_weekday = daily_weekday.reindex(
                        [d for d in day_order if d in daily_weekday.index]
                    )
                    daily_weekday.index = [day_names_pl[d] for d in daily_weekday.index]

                    col1, col2 = st.columns(2)

                    with col1:
                        fig_day_total = px.bar(
                            x=daily_weekday.index,
                            y=daily_weekday["sum"],
                            title="Łączne godziny per dzień tygodnia",
                            labels={"x": "Dzień", "y": "Godziny"},
                            color_discrete_sequence=["#1f77b4"],
                        )
                        fig_day_total.update_layout(height=350)
                        st.plotly_chart(fig_day_total, width="stretch")

                    with col2:
                        fig_day_avg = px.bar(
                            x=daily_weekday.index,
                            y=daily_weekday["mean"],
                            title="Średnio godzin per dzień tygodnia",
                            labels={"x": "Dzień", "y": "Średnia godzin"},
                            color_discrete_sequence=["#2ca02c"],
                        )
                        fig_day_avg.update_layout(height=350)
                        st.plotly_chart(fig_day_avg, width="stretch")

                    # Export per miesiąc
                    st.markdown("---")
                    st.subheader("📥 Pobierz raport miesiąca")

                    col1, col2 = st.columns(2)

                    with col1:
                        # CSV export
                        export_month_df = month_data[
                            [
                                "person",
                                "task",
                                "key",
                                "time_hours",
                                "creative_percent",
                                "creative_hours",
                            ]
                        ].copy()
                        export_month_df.columns = [
                            "Osoba",
                            "Zadanie",
                            "Klucz",
                            "Czas (h)",
                            "% Twórczości",
                            "Godziny twórcze",
                        ]

                        csv_data = export_month_df.to_csv(
                            index=False, encoding="utf-8-sig"
                        )
                        st.download_button(
                            label=f"📋 CSV - {selected_month}",
                            data=csv_data,
                            file_name=f"worklogs_{selected_month}_{start_date.strftime('%d')}-{end_date.strftime('%d')}.csv",
                            mime="text/csv",
                        )

                    with col2:
                        # Excel export - TAKI SAM FORMAT CO GŁÓWNY (2 arkusze)
                        buffer_month = io.BytesIO()

                        # Przygotuj dokumentację szczegółową
                        export_month_detailed = month_data[
                            ["person", "task", "key", "time_hours", "creative_percent", "creative_hours"]
                        ].copy()

                        # Dodaj kolumny takie same jak główny export
                        export_month_excel = pd.DataFrame()
                        export_month_excel["Osoba"] = export_month_detailed["person"]
                        export_month_excel["Zadanie"] = export_month_detailed["task"]
                        export_month_excel["Klucz"] = export_month_detailed["key"]
                        export_month_excel["Czas"] = export_month_detailed["time_hours"].apply(hours_to_hm_format)
                        export_month_excel["Czas (h)"] = export_month_detailed["time_hours"]
                        export_month_excel["Procent twórczości"] = export_month_detailed["creative_percent"]
                        export_month_excel["Godziny twórcze"] = export_month_detailed["creative_hours"].apply(
                            lambda x: f"{x:.1f}h" if pd.notna(x) else "Brak danych"
                        )
                        export_month_excel["Godziny twórcze (h)"] = export_month_detailed["creative_hours"]

                        with pd.ExcelWriter(buffer_month, engine="openpyxl") as writer:
                            # Arkusz 1: Szczegółowe dane
                            export_month_excel.to_excel(
                                writer, sheet_name="Worklogs", index=False
                            )

                            # Formatowanie arkusza szczegółowego
                            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                            worksheet_month = writer.sheets["Worklogs"]

                            # Dostosuj szerokość kolumn
                            column_widths_month = {
                                "A": 25,  # Osoba
                                "B": 50,  # Zadanie
                                "C": 15,  # Klucz
                                "D": 12,  # Czas
                                "E": 12,  # Czas (h)
                                "F": 18,  # % Twórczości
                                "G": 18,  # Godziny twórcze
                                "H": 18,  # Godziny twórcze (h)
                            }

                            for col, width in column_widths_month.items():
                                worksheet_month.column_dimensions[col].width = width

                            # Style formatowania
                            header_font_month = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                            header_fill_month = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
                            header_alignment_month = Alignment(horizontal="center", vertical="center")

                            thin_border_month = Border(
                                left=Side(style="thin", color="E0E0E0"),
                                right=Side(style="thin", color="E0E0E0"),
                                top=Side(style="thin", color="E0E0E0"),
                                bottom=Side(style="thin", color="E0E0E0"),
                            )

                            # Formatuj nagłówki
                            for cell in worksheet_month[1]:
                                cell.font = header_font_month
                                cell.fill = header_fill_month
                                cell.alignment = header_alignment_month
                                cell.border = thin_border_month

                            # Style dla danych
                            data_font_month = Font(name="Calibri", size=10)
                            data_alignment_left_month = Alignment(horizontal="left", vertical="center")
                            data_alignment_center_month = Alignment(horizontal="center", vertical="center")
                            data_alignment_right_month = Alignment(horizontal="right", vertical="center")

                            # Kolory dla procentów twórczości (takie same jak główny export)
                            color_red_month = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # 0-50%
                            color_yellow_month = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # 50-80%
                            color_green_month = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")  # 80-100%

                            # Formatuj dane i dodaj kolorowanie
                            for row_num in range(2, len(export_month_excel) + 2):
                                creative_percent_value = worksheet_month.cell(row=row_num, column=6).value

                                for col_num in range(1, len(export_month_excel.columns) + 1):
                                    cell = worksheet_month.cell(row=row_num, column=col_num)
                                    cell.font = data_font_month
                                    cell.border = thin_border_month

                                    if col_num in [1, 2]:  # Osoba, Zadanie
                                        cell.alignment = data_alignment_left_month
                                    elif col_num in [3]:  # Klucz
                                        cell.alignment = data_alignment_center_month
                                    elif col_num == 5:  # Czas (h)
                                        cell.alignment = data_alignment_right_month
                                        cell.number_format = "0.00"
                                    elif col_num == 6:  # % Twórczości
                                        cell.alignment = data_alignment_right_month
                                        cell.number_format = "0"
                                        # Kolorowanie
                                        if pd.notna(creative_percent_value) and creative_percent_value != "":
                                            try:
                                                percent = float(creative_percent_value)
                                                if percent <= 50:
                                                    cell.fill = color_red_month
                                                elif percent <= 80:
                                                    cell.fill = color_yellow_month
                                                else:
                                                    cell.fill = color_green_month
                                            except (ValueError, TypeError):
                                                pass
                                    elif col_num == 8:  # Godziny twórcze (h)
                                        cell.alignment = data_alignment_right_month
                                        cell.number_format = "0.00"
                                    else:  # Pozostałe kolumny
                                        cell.alignment = data_alignment_right_month

                            # Zamrożenie pierwszego wiersza
                            worksheet_month.freeze_panes = "A2"

                            # Filtr automatyczny
                            worksheet_month.auto_filter.ref = f"A1:{chr(65 + len(export_month_excel.columns) - 1)}{len(export_month_excel) + 1}"

                            # Arkusz 2: Podsumowanie per osoba
                            summary_month = month_data.groupby("person").agg({
                                "time_hours": "sum",
                                "creative_hours": "sum",
                                "creative_percent": lambda x: x.dropna().count(),
                            }).round(2)

                            # Oblicz czas TYLKO dla zadań z danymi o twórczości (worklogs)
                            time_hours_with_data_month = month_data[month_data["creative_percent"].notna()].groupby("person")["time_hours"].sum()
                            
                            # % twórczości ze ZGRUPOWANYCH GODZIN (gdzie mamy dane)
                            summary_month["creative_ratio"] = (
                                summary_month["creative_hours"] / time_hours_with_data_month * 100
                            ).round(1)

                            total_tasks_month = month_data.groupby("person").size()
                            summary_month["coverage"] = (
                                summary_month["creative_percent"] / total_tasks_month * 100
                            ).round(0)

                            summary_month = summary_month[["time_hours", "creative_hours", "creative_ratio", "coverage"]]
                            summary_month.columns = ["Łączne godziny", "Godziny twórcze", "% Pracy twórczej", "Pokrycie danymi"]
                            summary_month = summary_month.reset_index()

                            summary_month.to_excel(
                                writer, sheet_name="Podsumowanie", index=False
                            )

                            worksheet_summary_month = writer.sheets["Podsumowanie"]

                            # Szerokości kolumn
                            worksheet_summary_month.column_dimensions["A"].width = 25
                            worksheet_summary_month.column_dimensions["B"].width = 18
                            worksheet_summary_month.column_dimensions["C"].width = 18
                            worksheet_summary_month.column_dimensions["D"].width = 18
                            worksheet_summary_month.column_dimensions["E"].width = 18

                            # Formatuj nagłówki podsumowania
                            for cell in worksheet_summary_month[1]:
                                cell.font = header_font_month
                                cell.fill = header_fill_month
                                cell.alignment = header_alignment_month
                                cell.border = thin_border_month

                            # Formatuj dane podsumowania
                            for row_num in range(2, len(summary_month) + 2):
                                for col_num in range(1, len(summary_month.columns) + 1):
                                    cell = worksheet_summary_month.cell(row=row_num, column=col_num)
                                    cell.font = data_font_month
                                    cell.border = thin_border_month

                                    if col_num == 1:  # Osoba
                                        cell.alignment = data_alignment_left_month
                                    else:  # Liczby
                                        cell.alignment = data_alignment_right_month
                                        if col_num in [2, 3]:  # Godziny
                                            cell.number_format = "0.00"
                                        else:  # Procenty
                                            cell.number_format = "0.0"

                            worksheet_summary_month.freeze_panes = "A2"

                        buffer_month.seek(0)
                        st.download_button(
                            label=f"📊 Excel - {selected_month} (2 arkusze)",
                            data=buffer_month,
                            file_name=f"worklogs_{selected_month}_{start_date.strftime('%d')}-{end_date.strftime('%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )

        except FileNotFoundError:
            st.error("❌ Nie można znaleźć pliku. Spróbuj wgrać go ponownie.")
        except pd.errors.EmptyDataError:
            st.error("❌ Plik jest pusty lub uszkodzony.")
        except pd.errors.ParserError as e:
            st.error(f"❌ Błąd parsowania pliku Excel: {str(e)}")
            st.info("Upewnij się, że plik ma poprawną strukturę Excel (.xlsx)")
        except PermissionError:
            st.error(
                "❌ Brak dostępu do pliku. Upewnij się, że plik nie jest otwarty w innym programie."
            )
        except Exception as e:
            st.error(f"❌ Błąd podczas przetwarzania pliku: {str(e)}")
            st.info(
                "Sprawdź czy plik ma odpowiednią strukturę z kolumnami 'Level' i 'Users / Issues / Procent pracy twórczej'"
            )
            # Debug info for developers
            with st.expander("🐞 Szczegóły techniczne (dla developerów)"):
                st.code(str(e))
                import traceback

                st.code(traceback.format_exc())

    else:
        st.info("👈 Wgraj plik Excel w panelu bocznym aby rozpocząć analizę.")

        # Przykład struktury danych
        st.subheader("📋 Przykładowa struktura danych")
        example_data = pd.DataFrame(
            {
                "Level": [0, 1, 2, 1, 2, 0, 1, 2],
                "Users / Issues / Procent pracy twórczej": [
                    "Jan Kowalski",
                    "Implementacja modułu logowania",
                    "90",
                    "Testowanie aplikacji",
                    "No Procent pracy twórczej data",
                    "Anna Nowak",
                    "Projektowanie UI",
                    "100",
                ],
                "Key": ["", "PROJ-123", "", "PROJ-124", "", "", "PROJ-125", ""],
                "Total Time Spent": [
                    "",
                    "10:00",
                    "10:00",
                    "5:30",
                    "5:30",
                    "",
                    "8:15",
                    "8:15",
                ],
            }
        )
        st.dataframe(example_data, width="stretch")


if __name__ == "__main__":
    main()
